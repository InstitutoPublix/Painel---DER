"""
pipeline/assemble_der.py
Gera data/der_precomputed.json a partir das planilhas em data/.

Todas as leituras usam openpyxl read_only=True, data_only=True.

Divisão de papéis das fontes:
  • Contratos DOPSR1 por Regional.xlsx  → FONTE CANÔNICA de qualquer dado
    financeiro por SR (liquidado, empenhado, n_contratos, emergenciais).
    O campo "Região" na base 8398_e_8399 nunca teve regionalização real
    (~67% das linhas têm Município="9999999" e Região="4100"), por isso
    aquele arquivo foi removido do pipeline.
  • Condição da malha 2025.xlsx         → extensão (km) e liquidado/km por SR
  • IRI_e_FWD.xlsx                      → percentuais de condição por SR
"""

import io
import json
import os
import sys
from datetime import datetime
from collections import defaultdict

import openpyxl

# Garante UTF-8 no terminal Windows
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

# -------------------------------------------------------
# CAMINHOS
# -------------------------------------------------------
BASE = os.path.join(os.path.dirname(__file__), "..", "data")

F_CONTRATOS   = os.path.join(BASE, "Contratos DOPSR1 por Regional.xlsx")
F_MALHA       = os.path.join(BASE, "Condição da malha 2025.xlsx")
F_IRI_FWD     = os.path.join(BASE, "IRI_e_FWD.xlsx")
F_OUTPUT      = os.path.join(BASE, "der_precomputed.json")

# -------------------------------------------------------
# NORMALIZAÇÕES
# -------------------------------------------------------

def norm_sr(raw: str) -> str:
    """Converte variações do nome de SR para o formato canônico 'SR Xxxx'."""
    if raw is None:
        return ""
    s = str(raw).strip().upper()
    mapping = {
        "SR LESTE":          "SR Leste",
        "LESTE":             "SR Leste",
        "SR CAMPOS GERAIS":  "SR Campos Gerais",
        "CAMPOS GERAIS":     "SR Campos Gerais",
        "CAMPOS GERAIS/NORTE": "SR Campos Gerais",
        "CAMPOS GERAIS SUL": "SR Campos Gerais",
        "SR NORTE":          "SR Norte",
        "NORTE":             "SR Norte",
        "SR NOROESTE":       "SR Noroeste",
        "NOROESTE":          "SR Noroeste",
        "SR OESTE":          "SR Oeste",
        "OESTE":             "SR Oeste",
        "OESTE/SUDOESTE":    "SR Oeste",
        "SR SUDOESTE":       "SR Oeste",
    }
    return mapping.get(s, raw.strip())


def safe_float(v, default=0.0) -> float:
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def safe_int(v, default=0) -> int:
    try:
        return int(float(v)) if v is not None else default
    except (TypeError, ValueError):
        return default


# -------------------------------------------------------
# SEÇÃO 1 — CONTRATOS POR REGIONAL (fonte canônica de dados financeiros por SR)
# -------------------------------------------------------
print("▶ Seção 1: Contratos por regional…")

wb_con = openpyxl.load_workbook(F_CONTRATOS, read_only=True, data_only=True)
ws_con = wb_con["Empenho por contrato"]

rows_con = list(ws_con.iter_rows(values_only=True))
data_con = [r for r in rows_con[1:] if r[0] is not None and str(r[0]).startswith("CO")]

# Colunas: 0=Contrato, 1=Empenhado, 2=Liquidado(R$), 3=Pago, 4=SR, 5=TIPO CONTRATO
reg_acc = defaultdict(lambda: {
    "liquidado": 0.0, "empenhado": 0.0, "pago": 0.0,
    "n_contratos": 0, "emergencial": 0, "tipos": set()
})

for r in data_con:
    sr   = norm_sr(r[4])
    tipo = str(r[5]).strip().upper() if r[5] else ""
    reg_acc[sr]["liquidado"]   += safe_float(r[2])
    reg_acc[sr]["empenhado"]   += safe_float(r[1])
    reg_acc[sr]["pago"]        += safe_float(r[3])
    reg_acc[sr]["n_contratos"] += 1
    reg_acc[sr]["tipos"].add(tipo)
    if "EMERGENCIAL" in tipo:
        reg_acc[sr]["emergencial"] += 1

wb_con.close()

for sr, v in sorted(reg_acc.items()):
    print(f"   {sr}: liq={v['liquidado']:,.0f} n={v['n_contratos']} emg={v['emergencial']}")

# -------------------------------------------------------
# SEÇÃO 2 — CONDIÇÃO DA MALHA
# -------------------------------------------------------
print("▶ Seção 2: Condição da malha…")

wb_mal = openpyxl.load_workbook(F_MALHA, read_only=True, data_only=True)
ws_mal = wb_mal["Malha (km)"]

rows_mal = list(ws_mal.iter_rows(values_only=True))
# Linha 0 = cabeçalho; dados de 1 em diante
# Colunas: 0=SR, 1=Ruim, 2=Péssimo, 3=Regular, 4=Boa, 5=Muito Boa, 6=Liquidado
malha_data = {}

for r in rows_mal[1:]:
    if r[0] is None:
        continue
    sr        = norm_sr(r[0])
    ruim      = safe_float(r[1])
    pessimo   = safe_float(r[2])
    regular   = safe_float(r[3])
    boa       = safe_float(r[4])
    muito_boa = safe_float(r[5])
    liq_sr    = safe_float(r[6])

    km_total = ruim + pessimo + regular + boa + muito_boa
    malha_data[sr] = {
        "km_ruim":      ruim,
        "km_pessimo":   pessimo,
        "km_regular":   regular,
        "km_boa":       boa,
        "km_muito_boa": muito_boa,
        "km_total":     round(km_total, 2),
        "pct_ruim_pessimo":  round((ruim + pessimo) / km_total * 100, 2) if km_total > 0 else 0,
        "pct_regular":       round(regular / km_total * 100, 2) if km_total > 0 else 0,
        "pct_bom_muito_bom": round((boa + muito_boa) / km_total * 100, 2) if km_total > 0 else 0,
        "liquidado_malha":   round(liq_sr, 2),
        "liquidado_por_km":  round(liq_sr / km_total, 2) if km_total > 0 else 0,
    }

wb_mal.close()

for sr, v in sorted(malha_data.items()):
    print(f"   {sr}: km_total={v['km_total']:.0f} pct_crit={v['pct_ruim_pessimo']:.1f}% lkm={v['liquidado_por_km']:,.0f}")

# -------------------------------------------------------
# SEÇÃO 3 — IRI e FWD
# -------------------------------------------------------
print("▶ Seção 3: IRI e FWD…")

wb_iri = openpyxl.load_workbook(F_IRI_FWD, read_only=True, data_only=True)

def parse_iri_fwd_sheet(ws):
    """
    Layout da aba:
      Linha 0: título
      Linha 1: cabeçalho — col0='Condição', cols 1-5 = nomes dos SRs
      Linhas 2-6: Ruim, Péssimo, Regular, Bom, Muito Bom (valores % por SR)
    Retorna dict: { sr_name: {ruim, pessimo, regular, bom, muito_bom} }
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 7:
        return {}

    header = rows[1]
    srs = [norm_sr(header[i]) for i in range(1, 6) if header[i] is not None]

    categorias = ["ruim", "pessimo", "regular", "bom", "muito_bom"]
    result = {sr: {} for sr in srs}

    for i, cat in enumerate(categorias):
        row = rows[2 + i]
        for j, sr in enumerate(srs):
            result[sr][cat] = round(safe_float(row[j + 1]), 2)

    return result


iri_data = parse_iri_fwd_sheet(wb_iri["IRI"])
fwd_data = parse_iri_fwd_sheet(wb_iri["FWD"])
wb_iri.close()

for sr in sorted(iri_data):
    d = iri_data[sr]
    pct_crit = d["ruim"] + d["pessimo"]
    print(f"   IRI {sr}: crítico={pct_crit:.1f}%")

# -------------------------------------------------------
# MONTAGEM DO JSON FINAL
# -------------------------------------------------------
print("▶ Montando JSON final…")

todas_srs = sorted(set(
    list(reg_acc.keys()) +
    list(malha_data.keys()) +
    list(iri_data.keys())
))

regionais_json = {}
for sr in todas_srs:
    con = reg_acc.get(sr, {})
    mal = malha_data.get(sr, {})
    iri = iri_data.get(sr, {})
    fwd = fwd_data.get(sr, {})

    regionais_json[sr] = {
        # Dados financeiros (fonte: Contratos DOPSR1 por Regional.xlsx)
        "liquidado":   round(con.get("liquidado", 0.0), 2),
        "empenhado":   round(con.get("empenhado", 0.0), 2),
        "n_contratos": con.get("n_contratos", 0),
        "emergencial": con.get("emergencial", 0),
        "tipos_contrato": sorted(con.get("tipos", set())),

        # Dados de malha (km) — fonte: Condição da malha 2025.xlsx
        "km_ruim":      mal.get("km_ruim", 0),
        "km_pessimo":   mal.get("km_pessimo", 0),
        "km_regular":   mal.get("km_regular", 0),
        "km_boa":       mal.get("km_boa", 0),
        "km_muito_boa": mal.get("km_muito_boa", 0),
        "km_total":     mal.get("km_total", 0),
        "pct_ruim_pessimo":  mal.get("pct_ruim_pessimo", 0),
        "pct_regular":       mal.get("pct_regular", 0),
        "pct_bom_muito_bom": mal.get("pct_bom_muito_bom", 0),
        "liquidado_por_km":  mal.get("liquidado_por_km", 0),

        # IRI — % do total inspecionado (fonte: IRI_e_FWD.xlsx)
        "iri": iri if iri else {},

        # FWD — % do total inspecionado (fonte: IRI_e_FWD.xlsx)
        "fwd": fwd if fwd else {},
    }

output = {
    "generated": datetime.now().isoformat(timespec="seconds"),
    "regionais": regionais_json,
}

with open(F_OUTPUT, "w", encoding="utf-8") as f:
    json.dump(output, f, ensure_ascii=False, indent=2)

liq_total = sum(v["liquidado"] for v in regionais_json.values())

print()
print("=" * 55)
print(f"✅ JSON gerado: {F_OUTPUT}")
print(f"   Liquidado total (Contratos DOPSR1): R$ {liq_total:>14,.0f}")
print(f"   Regionais:  {len(regionais_json)}")
print("=" * 55)
