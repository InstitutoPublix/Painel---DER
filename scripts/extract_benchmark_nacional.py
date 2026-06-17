"""
extract_benchmark_nacional.py
==============================
Extrai os indicadores de condição de rodovias estaduais (CNT / SEFA)
da planilha INDICADORES SEFA_GERAL.xlsx e gera dashboard/data/benchmark_nacional.json.

Para re-executar quando a planilha for atualizada:
    python scripts/extract_benchmark_nacional.py

Execute sempre a partir da raiz do projeto (diretório que contém dashboard/ e data/).
O arquivo de entrada é lido de   data/INDICADORES SEFA_GERAL.xlsx
O arquivo de saída é gravado em  dashboard/data/benchmark_nacional.json
"""

import json
import os
import sys
from datetime import date
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("Dependência ausente: instale com  pip install openpyxl")

# ── Caminhos ────────────────────────────────────────────────────────────────
ROOT   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX   = os.path.join(ROOT, "data", "INDICADORES SEFA_GERAL.xlsx")
OUT    = os.path.join(ROOT, "dashboard", "data", "benchmark_nacional.json")

if not os.path.exists(XLSX):
    sys.exit(f"Planilha não encontrada: {XLSX}")

print(f"Abrindo: {XLSX}")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# ── Nomes das colunas que precisamos ────────────────────────────────────────
COL_ESTADO = "Estado"
COL_UF     = "UF"
COL_ANO    = "Ano"
COL_OTIM   = "9.7.1.Log_EstRodovOtim_UF(%)"
COL_BOA    = "9.7.2.Log_EstRodovBom_UF(%)"

def find_header(ws, target_col, max_rows=5):
    """Retorna (row_index_0based, list_header) da primeira linha que contenha target_col."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True)):
        if target_col in row:
            return i, list(row)
    return None, None

# ── Aba ESTADOS (cabeçalho na linha 2) ──────────────────────────────────────
ws_est = wb["ESTADOS"]
_, hdr_est = find_header(ws_est, COL_UF)
if hdr_est is None:
    sys.exit("Coluna 'UF' não encontrada na aba ESTADOS — verifique o arquivo.")

ci_estado = hdr_est.index(COL_ESTADO)
ci_uf     = hdr_est.index(COL_UF)
ci_ano    = hdr_est.index(COL_ANO)
ci_otim   = hdr_est.index(COL_OTIM)
ci_boa    = hdr_est.index(COL_BOA)

# Monta nome canônico de cada UF (primeira ocorrência)
uf_nome = {}
# Agrega pct_boa_otima por (UF, ano) — a planilha pode ter mais de uma linha por UF/ano;
# neste caso soma-se apenas se ambos os valores existirem
uf_ano_vals = defaultdict(list)

# iter_rows a partir da linha após o cabeçalho
skip = True
for row in ws_est.iter_rows(min_row=1, values_only=True):
    if skip:
        if row[ci_uf] == "UF":   # pula a linha de cabeçalho
            skip = False
        continue
    uf   = row[ci_uf]
    ano  = row[ci_ano]
    nome = row[ci_estado]
    if not uf or not ano:
        continue
    if uf not in uf_nome and nome:
        uf_nome[uf] = nome
    otim = row[ci_otim]
    boa  = row[ci_boa]
    if otim is None and boa is None:
        continue
    total = round((otim or 0.0) + (boa or 0.0), 2)
    uf_ano_vals[(uf, int(ano))].append(total)

# Consolida: se houver mais de um valor por UF/ano, usa a média (não esperado, mas robusto)
estados = {}
for (uf, ano), vals in uf_ano_vals.items():
    if uf not in estados:
        estados[uf] = {"nome": uf_nome.get(uf, uf), "serie": {}}
    estados[uf]["serie"][str(ano)] = round(sum(vals) / len(vals), 2)

# ── Aba BRASIL (cabeçalho na linha 1) ───────────────────────────────────────
ws_br = wb["BRASIL"]
_, hdr_br = find_header(ws_br, COL_ANO)
if hdr_br is None:
    sys.exit("Coluna 'Ano' não encontrada na aba BRASIL — verifique o arquivo.")

ci_ano_br  = hdr_br.index(COL_ANO)
ci_otim_br = hdr_br.index(COL_OTIM)
ci_boa_br  = hdr_br.index(COL_BOA)

brasil = {}
skip2 = True
for row in ws_br.iter_rows(min_row=1, values_only=True):
    if skip2:
        if row[ci_ano_br] == "Ano":
            skip2 = False
        continue
    ano  = row[ci_ano_br]
    otim = row[ci_otim_br]
    boa  = row[ci_boa_br]
    if not ano:
        continue
    if otim is None and boa is None:
        continue
    brasil[str(int(ano))] = round((otim or 0.0) + (boa or 0.0), 2)

# ── Anos da série (união de todos os anos com dado em qualquer UF ou Brasil) ─
all_anos = sorted({
    int(a)
    for est in estados.values()
    for a in est["serie"]
} | {int(a) for a in brasil})

# ── Saída ────────────────────────────────────────────────────────────────────
payload = {
    "estados":    estados,
    "brasil":     brasil,
    "anos":       all_anos,
    "gerado_em":  str(date.today()),
    "fonte":      (
        "INDICADORES SEFA_GERAL.xlsx, abas ESTADOS e BRASIL — "
        "colunas localizadas por nome de cabecalho: "
        f"{COL_OTIM} + {COL_BOA}"
    )
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(payload, f, ensure_ascii=False, indent=2)

n_uf   = len(estados)
n_anos = len(all_anos)
print(f"OK — {n_uf} estados / {n_anos} anos extraidos")
print(f"Anos: {all_anos}")
print(f"Arquivo gerado: {OUT}")

# Resumo de validação rápida
print("\nValidacao (PR / SP / AC / Brasil):")
for uf, ano in [("PR","2024"),("PR","2025"),("SP","2025"),("AC","2025")]:
    val = estados.get(uf,{}).get("serie",{}).get(ano,"AUSENTE")
    print(f"  {uf} {ano}: {val}%")
for ano in ("2024","2025"):
    print(f"  Brasil {ano}: {brasil.get(ano,'AUSENTE')}%")
